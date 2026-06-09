# Reads source → calculates → writes to Rewards.xlsx on SharePoint

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import io
import json

# Load credentials from .env file
try:
    from dotenv import dotenv_values
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "python-dotenv"])
    from dotenv import dotenv_values

config = dotenv_values(r"D:\SHAREPOINT\.env")
USERNAME = config.get("SHAREPOINT_USERNAME")
PASSWORD = config.get("SHAREPOINT_PASSWORD")
TENANT = config.get("SHAREPOINT_TENANT")

# SharePoint Site and Drive IDs (from MCP discovery)
SOURCE_DRIVE_ID = "b!_Oj5AOOCqUa-6fnpgxmwM4Tmz3IIfOZIhM-bF3vfV8Q7o8oZ3WyrQ4ILTnuUDgHw"
SOURCE_ITEM_ID = "01EUH7IGAHNG3EYW2JJ5C37HVRDHKNUFDB"

OUTPUT_DRIVE_ID = "b!ncIFrojL106h53P8D_qyinjcKBnN9CFGrFJ26Ac7DPP-kaV-qe0HS73NEabHbqpx"
OUTPUT_ITEM_ID = "01ZAECLIUEFXFIUSMHQNAZRWSUI2DSRS5G"

EMPLOYEES = [
    "Pradip Ray", "Srijan Banerjee", "Anurima Nath", "Madhuparna Naskar",
    "Atreyee Majumder", "Pratim Rath", "Rahul Paul",
    "Monirul Molla", "Debnath Paul", "Debrup Mondal"
]

POINT_COLUMNS = ["Punctuality", "L&D", "Fluency Compliance", "Innovation", "Extraordinary Performance"]

def get_access_token():
    """Get access token using client credentials or device flow"""
    try:
        from msal import PublicClientApplication

        CLIENT_ID = "04b07795-8ddb-461a-bbee-02f9e1bf7b46"
        AUTHORITY = f"https://login.microsoftonline.com/{TENANT}"
        SCOPES = ["https://graph.microsoft.com/.default"]

        app = PublicClientApplication(CLIENT_ID, authority=AUTHORITY)

        # Try to get token using username/password (resource owner password credentials flow)
        result = app.acquire_token_by_username_password(
            username=USERNAME,
            password=PASSWORD,
            scopes=SCOPES
        )

        if "access_token" in result:
            return result["access_token"]
        else:
            print(f"Error: {result.get('error_description', 'Unknown error')}")
            return None
    except Exception as e:
        print(f"Auth error: {e}")
        return None

def download_file(drive_id, item_id, token):
    """Download file from SharePoint using Graph API"""
    graph_url = "https://graph.microsoft.com/v1.0"
    headers = {"Authorization": f"Bearer {token}"}

    # Get file content directly using drive ID and item ID
    file_url = f"{graph_url}/drives/{drive_id}/items/{item_id}/content"
    response = requests.get(file_url, headers=headers)

    if response.status_code == 200:
        return io.BytesIO(response.content)
    else:
        print(f"Error downloading file: {response.status_code} - {response.text}")
        return None

def upload_file(drive_id, item_id, file_content, token):
    """Upload file to SharePoint using Graph API with upload session"""
    graph_url = "https://graph.microsoft.com/v1.0"

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    # Create upload session
    upload_url = f"{graph_url}/drives/{drive_id}/items/{item_id}/createUploadSession"

    response = requests.post(upload_url, headers=headers, json={
        "item": {
            "@microsoft.graph.conflictBehavior": "replace"
        }
    })

    if response.status_code != 200:
        print(f"Error creating upload session: {response.status_code} - {response.text}")
        # Try direct upload as fallback
        return upload_file_direct(drive_id, item_id, file_content, token)

    upload_session = response.json()["uploadUrl"]

    # Upload file using the session
    file_size = len(file_content)
    upload_headers = {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Length": str(file_size),
        "Content-Range": f"bytes 0-{file_size-1}/{file_size}"
    }

    response = requests.put(upload_session, data=file_content, headers=upload_headers)

    if response.status_code in [200, 201]:
        return True
    else:
        print(f"Error uploading file via session: {response.status_code} - {response.text}")
        return False

def upload_file_direct(drive_id, item_id, file_content, token):
    """Direct file upload to SharePoint"""
    graph_url = "https://graph.microsoft.com/v1.0"

    # Simple PUT request without special headers
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

    file_url = f"{graph_url}/drives/{drive_id}/items/{item_id}/content"
    response = requests.put(file_url, data=file_content, headers=headers)

    if response.status_code in [200, 201]:
        print("  Upload successful!")
        return True
    else:
        print(f"Error in direct upload: {response.status_code}")
        if response.text:
            print(f"Response: {response.text[:300]}")
        return False

# Get last Monday to Sunday
today = datetime.today()
last_monday = today - timedelta(days=today.weekday() + 7)
last_sunday = last_monday + timedelta(days=6)
week_number = last_monday.isocalendar()[1]
week_label = f"Week #{week_number} - {last_monday.strftime('%b %d')} - {last_sunday.strftime('%b %d')}"
print(f"Processing: {week_label}")

# Get access token
print("Authenticating...")
token = get_access_token()
if not token:
    print("Authentication failed. Please check your credentials.")
    exit(1)

print("Downloading source file...")
src_buffer = download_file(SOURCE_DRIVE_ID, SOURCE_ITEM_ID, token)
if not src_buffer:
    print("ERROR: Could not download source file")
    exit(1)

src_wb = load_workbook(src_buffer)

# Find the daily data sheet
src_ws = None
for sheet in src_wb.sheetnames:
    ws = src_wb[sheet]
    headers = [str(ws.cell(1, c).value).strip() for c in range(1, ws.max_column+1)]
    if "Date" in headers and "Name" in headers and "Punctuality" in headers:
        src_ws = ws
        break

if not src_ws:
    print("ERROR: Could not find daily data sheet")
    exit()

# Parse headers
headers = [str(src_ws.cell(1, c).value).strip() for c in range(1, src_ws.max_column+1)]
col = {h: i+1 for i, h in enumerate(headers)}

# Calculate points per employee for last week
results = {emp: 0 for emp in EMPLOYEES}
for row in range(2, src_ws.max_row+1):
    raw_date = src_ws.cell(row, col.get("Date", 1)).value
    if not raw_date:
        continue
    if isinstance(raw_date, str):
        try:
            raw_date = datetime.strptime(raw_date, "%m/%d/%Y")
        except:
            continue
    if last_monday.date() <= raw_date.date() <= last_sunday.date():
        name = str(src_ws.cell(row, col.get("Name", 1)).value).strip()
        if name in results:
            for pc in POINT_COLUMNS:
                if pc in col:
                    val = src_ws.cell(row, col[pc]).value
                    try:
                        results[name] += float(val) if val else 0
                    except:
                        pass

print("Downloading output file...")
out_buffer_read = download_file(OUTPUT_DRIVE_ID, OUTPUT_ITEM_ID, token)
if not out_buffer_read:
    print("ERROR: Could not download output file")
    exit(1)

out_wb = load_workbook(out_buffer_read)

# Use fixed sheet name "Weekly Rewards"
SHEET_NAME = "Weekly Rewards"
if SHEET_NAME in out_wb.sheetnames:
    ws_new = out_wb[SHEET_NAME]
    # Clear existing data
    ws_new.delete_rows(1, ws_new.max_row)
    print(f"Cleared existing '{SHEET_NAME}' sheet")
else:
    ws_new = out_wb.create_sheet(title=SHEET_NAME)
    print(f"Created new '{SHEET_NAME}' sheet")

# Write header with week label
ws_new.append([week_label])
ws_new.append([])  # Empty row
ws_new.append(["Name", "Points", "Amount"])

# Write employee data
total_pts = 0
for emp in EMPLOYEES:
    pts = int(results[emp])
    amt = pts * 10
    total_pts += pts
    ws_new.append([emp, pts, f"₹{amt}"])

# Format the sheet
ws_new['A1'].font = Font(bold=True)
for col in range(1, 4):
    ws_new.cell(row=3, column=col).alignment = Alignment(horizontal='center')

print(f"'{SHEET_NAME}' sheet updated successfully")

# Save to local file first
print("Saving updated file locally...")
import time
local_file_path = r"d:\sharepoint\Rewards.xlsx"
out_buffer_write = io.BytesIO()
out_wb.save(out_buffer_write)
out_buffer_write.seek(0)

# Write to local file
with open(local_file_path, 'wb') as f:
    f.write(out_buffer_write.getvalue())
print(f"Saved locally to: {local_file_path}")

# Wait before uploading
print("Waiting 5 seconds before uploading to SharePoint...")
time.sleep(5)

# Upload to SharePoint with force-overwrite headers
print("Uploading updated file to SharePoint...")
out_buffer_write.seek(0)

if upload_file_direct(OUTPUT_DRIVE_ID, OUTPUT_ITEM_ID, out_buffer_write.getvalue(), token):
    print("SUCCESS! Weekly report written to Rewards.xlsx on SharePoint")

    # Clean up local files after successful upload
    import os
    local_files = [
        r"D:\SHAREPOINT\Rewards.xlsx",
        r"D:\SHAREPOINT\Book2.xlsx",
        r"D:\SHAREPOINT\Book2_backup_20260608_164447.xlsx"
    ]
    for f in local_files:
        if os.path.exists(f):
            os.remove(f)
            print(f"Deleted local temp file: {f}")
else:
    print("Upload to SharePoint failed (file may be temporarily locked)")
    print(f"Local copy is ready at: {local_file_path}")
    print("\nYou can manually upload this file to SharePoint when the lock clears")
