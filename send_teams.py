# Reads Rewards.xlsx → sends to Teams

import requests
from openpyxl import load_workbook
from datetime import datetime
import io
import json

# Load credentials from .env file
try:
    from dotenv import dotenv_values
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "python-dotenv"])
    from dotenv import dotenv_values

config = dotenv_values(r"D:\SHAREPOINT\.env")
USERNAME = config.get("SHAREPOINT_USERNAME")
PASSWORD = config.get("SHAREPOINT_PASSWORD")
TENANT = config.get("SHAREPOINT_TENANT")

# SharePoint Site and Drive IDs (from MCP discovery)
OUTPUT_DRIVE_ID = "b!ncIFrojL106h53P8D_qyinjcKBnN9CFGrFJ26Ac7DPP-kaV-qe0HS73NEabHbqpx"
OUTPUT_ITEM_ID = "01ZAECLIUEFXFIUSMHQNAZRWSUI2DSRS5G"

# Teams Webhook URL
WEBHOOK_URL = "https://defaultabe0ba584a8e4ee9985a85449c16df.58.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/7a93c82bdb7b4095894f25309cbe4673/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=NymSHVyWZFGo_y0NQLWsEeapIiWl7f0L44B8z0zYAW4"

def get_access_token():
    """Get access token using delegated auth (same as weekly_report.py)"""
    try:
        from msal import PublicClientApplication

        CLIENT_ID = "04b07795-8ddb-461a-bbee-02f9e1bf7b46"
        AUTHORITY = f"https://login.microsoftonline.com/{TENANT}"
        # Use .default scope which should include Chat permissions for delegated auth
        SCOPES = ["https://graph.microsoft.com/.default"]

        app = PublicClientApplication(CLIENT_ID, authority=AUTHORITY)
        result = app.acquire_token_by_username_password(
            username=USERNAME,
            password=PASSWORD,
            scopes=SCOPES
        )

        if "access_token" in result:
            return result["access_token"]
        else:
            print(f"Error: {result.get('error_description', 'Unknown error')}")
            return None
    except Exception as e:
        print(f"Auth error: {e}")
        return None

def download_file(drive_id, item_id, token):
    """Download file from SharePoint using Graph API"""
    graph_url = "https://graph.microsoft.com/v1.0"
    headers = {"Authorization": f"Bearer {token}"}

    file_url = f"{graph_url}/drives/{drive_id}/items/{item_id}/content"
    response = requests.get(file_url, headers=headers)

    if response.status_code == 200:
        return io.BytesIO(response.content)
    else:
        print(f"Error downloading file: {response.status_code} - {response.text}")
        return None

def read_weekly_rewards(buffer):
    """Read data from Weekly Rewards sheet"""
    try:
        wb = load_workbook(buffer)
        if "Weekly Rewards" not in wb.sheetnames:
            print("ERROR: 'Weekly Rewards' sheet not found")
            return None

        ws = wb["Weekly Rewards"]

        # Extract week label from Row 1
        week_label = ws['A1'].value
        if not week_label:
            print("ERROR: Could not find week label in Row 1")
            return None

        # Extract headers from Row 3
        headers = [ws.cell(row=3, column=c).value for c in range(1, 4)]

        # Extract employee data starting from Row 4
        employees = []
        row = 4
        while True:
            name = ws.cell(row=row, column=1).value
            if not name:
                break
            points = ws.cell(row=row, column=2).value
            amount = ws.cell(row=row, column=3).value
            employees.append({
                "name": name,
                "points": points,
                "amount": amount
            })
            row += 1

        return {
            "week_label": week_label,
            "headers": headers,
            "employees": employees
        }
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def format_teams_message(data):
    """Format data as Adaptive Card"""
    week_label = data["week_label"]
    employees = data["employees"]

    # Build adaptive card body
    body = [
        {
            "type": "TextBlock",
            "text": "Weekly Performance Report",
            "weight": "Bolder",
            "size": "Large",
            "color": "Accent"
        },
        {
            "type": "TextBlock",
            "text": week_label,
            "weight": "Bolder",
            "size": "Medium",
            "spacing": "None"
        },
        {
            "type": "ColumnSet",
            "separator": True,
            "columns": [
                {"type": "Column", "width": "stretch", "items": [{"type": "TextBlock", "text": "Name", "weight": "Bolder"}]},
                {"type": "Column", "width": "100px", "items": [{"type": "TextBlock", "text": "Points", "weight": "Bolder", "horizontalAlignment": "Center"}]},
                {"type": "Column", "width": "100px", "items": [{"type": "TextBlock", "text": "Amount", "weight": "Bolder", "horizontalAlignment": "Right"}]}
            ]
        }
    ]

    # Add employee rows
    for idx, emp in enumerate(employees):
        name = emp["name"]
        points = emp["points"] if emp["points"] is not None else 0
        amount = emp["amount"] if emp["amount"] else "0"

        row = {
            "type": "ColumnSet",
            "separator": True,
            "style": "default",
            "columns": [
                {"type": "Column", "width": "stretch", "items": [{"type": "TextBlock", "text": str(name)}]},
                {"type": "Column", "width": "100px", "items": [{"type": "TextBlock", "text": str(points), "horizontalAlignment": "Center"}]},
                {"type": "Column", "width": "100px", "items": [{"type": "TextBlock", "text": str(amount), "horizontalAlignment": "Right"}]}
            ]
        }

        body.append(row)

    # Build complete adaptive card
    card = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": body
    }

    return card

def get_user_chats(token):
    """Get list of user's chats"""
    graph_url = "https://graph.microsoft.com/v1.0"
    headers = {"Authorization": f"Bearer {token}"}

    url = f"{graph_url}/me/chats"
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        return response.json().get("value", [])
    else:
        print(f"Error getting chats: {response.status_code}")
        if response.text:
            print(f"Response: {response.text[:300]}")
        return None

def send_teams_webhook_message(webhook_url, adaptive_card):
    """Send Adaptive Card message to Teams using webhook"""
    headers = {"Content-Type": "application/json"}

    # Wrap adaptive card in message format
    payload = {
        "type": "message",
        "attachments": [
            {
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": adaptive_card
            }
        ]
    }

    response = requests.post(webhook_url, headers=headers, json=payload)

    if response.status_code in [200, 201, 202]:
        return True
    else:
        print(f"Error sending message: {response.status_code}")
        if response.text:
            print(f"Response: {response.text[:300]}")
        return False

# MAIN EXECUTION
print("Step 1: Authenticating to SharePoint...")
token = get_access_token()
if not token:
    print("Authentication failed. Please check your credentials.")
    exit(1)

print("Step 2: Downloading Rewards.xlsx from SharePoint...")
file_buffer = download_file(OUTPUT_DRIVE_ID, OUTPUT_ITEM_ID, token)
if not file_buffer:
    print("ERROR: Could not download Rewards.xlsx")
    exit(1)

print("Step 3: Reading Weekly Rewards sheet...")
data = read_weekly_rewards(file_buffer)
if not data:
    print("ERROR: Could not read Weekly Rewards sheet")
    exit(1)

print(f"Found data for: {data['week_label']}")
print(f"Found {len(data['employees'])} employees")

print("Step 4: Formatting Teams message...")
message = format_teams_message(data)

print("Step 5: Sending message to Teams via Webhook...")
if send_teams_webhook_message(WEBHOOK_URL, message):
    print("[SUCCESS] Weekly report sent to Teams successfully!")
else:
    print("ERROR: Failed to send message to Teams")
    exit(1)
