# Downloads source file → sends performance report to Teams

import requests
from openpyxl import load_workbook
from collections import defaultdict
import os
import tempfile
import time
import json
from datetime import datetime, timedelta
from dotenv import load_dotenv

# Load .env file if it exists
load_dotenv()

# Load credentials from environment variables (Azure App Registration)
AZURE_CLIENT_ID = os.environ.get("AZURE_CLIENT_ID")
AZURE_TENANT_ID = os.environ.get("AZURE_TENANT_ID")
AZURE_CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET")

# Teams Webhook URL — override via TEAMS_WEBHOOK_URL env var (e.g. GitHub Actions secret);
# falls back to the hardcoded default so local runs work without extra setup.
WEBHOOK_URL = os.environ.get("TEAMS_WEBHOOK_URL") or "https://defaultabe0ba584a8e4ee9985a85449c16df.58.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/7a93c82bdb7b4095894f25309cbe4673/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=NymSHVyWZFGo_y0NQLWsEeapIiWl7f0L44B8z0zYAW4"

# File paths — uses the OS temp directory so it works on both Windows (local) and
# Linux (GitHub Actions runner).
TEMP_FILE = os.path.join(tempfile.gettempdir(), "temp_source.xlsx")

# SharePoint API details
SHAREPOINT_DRIVE_ID = "b!_Oj5AOOCqUa-6fnpgxmwM4Tmz3IIfOZIhM-bF3vfV8Q7o8oZ3WyrQ4ILTnuUDgHw"
SHAREPOINT_ITEM_ID = "01EUH7IGAHNG3EYW2JJ5C37HVRDHKNUFDB"

SNAPSHOT_FILE = "last_known_yesterday.json"


def get_access_token():
    try:
        from msal import ConfidentialClientApplication
        authority = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
        app = ConfidentialClientApplication(
            AZURE_CLIENT_ID,
            client_credential=AZURE_CLIENT_SECRET,
            authority=authority
        )
        result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        if "access_token" in result:
            return result["access_token"]
        else:
            print(f"Note: Authentication failed - {result.get('error_description', result.get('error', 'Unknown error'))}")
            return None
    except Exception as e:
        print(f"Note: Could not authenticate ({type(e).__name__})")
        return None


def download_from_sharepoint_api(token):
    try:
        headers = {"Authorization": f"Bearer {token}"}
        file_url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{SHAREPOINT_ITEM_ID}/content"
        response = requests.get(file_url, headers=headers, timeout=30)
        if response.status_code == 200:
            temp_dir = os.path.dirname(TEMP_FILE)
            if temp_dir:
                os.makedirs(temp_dir, exist_ok=True)
            with open(TEMP_FILE, 'wb') as f:
                f.write(response.content)
            print("[OK] Source file downloaded successfully")
            return True
        print(f"[ERROR] SharePoint API returned status {response.status_code}")
        return False
    except Exception as e:
        print(f"[ERROR] Could not download via API: {e}")
        return False


def download_source_file():
    print("Attempting to download from public SharePoint link...")
    url = "https://cloudaiorg.sharepoint.com/:x:/r/sites/StrictEmployeePerformance/_layouts/15/Doc.aspx?sourcedoc=%7B4CB66907-495B-454F-BF9E-B119D4DA1461%7D&file=Strict%20Employee%20Performance%20Analysis.xlsx&fromShare=true&action=default&mobileredirect=true&download=1"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream",
        "Accept-Language": "en-US,en;q=0.9"
    }
    try:
        session = requests.Session()
        response = session.get(url, headers=headers, allow_redirects=True, timeout=30)
        response.raise_for_status()
        if response.headers.get('content-type', '').startswith('text/html') or b'<!DOCTYPE' in response.content[:100]:
            print("Note: Public link requires authentication. Using environment variable credentials...")
        else:
            temp_dir = os.path.dirname(TEMP_FILE)
            if temp_dir:
                os.makedirs(temp_dir, exist_ok=True)
            with open(TEMP_FILE, 'wb') as f:
                f.write(response.content)
            print("[OK] Source file downloaded successfully")
            return True
    except Exception as e:
        print(f"Note: Public link failed ({type(e).__name__})")

    if AZURE_CLIENT_ID and AZURE_TENANT_ID and AZURE_CLIENT_SECRET:
        for attempt in range(1, 4):
            print(f"Attempting authenticated download via Microsoft Graph API (attempt {attempt}/3)...")
            token = get_access_token()
            if token and download_from_sharepoint_api(token):
                return True
            if attempt < 3:
                print("Authenticated download failed, retrying in 30 seconds...")
                time.sleep(30)

    print("[ERROR] Failed to download source file from all methods.")
    return False


# Header names (case-insensitive) that are never point/category columns,
# beyond Date/Name/Notes which are handled separately.
NON_POINT_COLUMNS = {"index"}


def _detect_columns(ws):
    """Return (date_col, name_col, category_cols, point_cols, notes_col) from header row.

    All indices are 1-based. date_col, name_col, and notes_col are None if not found.
    category_cols maps category name → column index for the five bonus categories.
    point_cols lists every non-date, non-name, non-index, non-free-text scoring column.
    notes_col is the free-text "Notes" column, kept separate from the numeric columns.
    """
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    date_col = name_col = notes_col = None
    point_cols = []
    category_cols = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        hl = str(h).strip().lower()
        if hl == "date":
            date_col = i + 1
        elif hl == "name":
            name_col = i + 1
        elif hl == "notes":
            notes_col = i + 1
        elif hl not in NON_POINT_COLUMNS:
            point_cols.append(i + 1)
            h_clean = str(h).strip()
            category_cols[h_clean] = i + 1
    return date_col, name_col, category_cols, point_cols, notes_col


def _safe_int(value, context=""):
    """Convert value to int, tolerating stray non-numeric cell values.

    Returns 0 for None/empty. On a non-numeric value, prints a warning
    (including context, e.g. row/employee/category) and returns 0 instead
    of crashing the whole report.
    """
    if not value:
        return 0
    try:
        return int(value)
    except (ValueError, TypeError):
        print(f"[WARNING] Non-numeric value in point column ({context}): {value!r} — treated as 0")
        return 0


def get_real_yesterday_data():
    """Read per-employee category data for actual calendar yesterday.

    Always uses today - 1 day (never weekday-adjusted).
    Returns {employee_name: {category: int}} for rows matching yesterday's date.
    Last occurrence wins when duplicate (name, date) rows exist.
    """
    yesterday_date = datetime.now().date() - timedelta(days=1)
    wb = load_workbook(TEMP_FILE)
    if "Daily Performance Bonus" in wb.sheetnames:
        ws = wb["Daily Performance Bonus"]
    else:
        ws = max(wb.worksheets, key=lambda s: s.max_row or 0)

    date_col, name_col, category_cols, _, _ = _detect_columns(ws)
    if not name_col or not date_col:
        return {}

    result = {}
    for row in range(2, ws.max_row + 1):
        name_val = ws.cell(row=row, column=name_col).value
        date_val = ws.cell(row=row, column=date_col).value
        if not name_val or not date_val:
            continue
        row_date = date_val.date() if hasattr(date_val, "date") else None
        if not row_date or row_date != yesterday_date:
            continue
        name = str(name_val).strip()
        result[name] = {
            cat_name: _safe_int(
                ws.cell(row=row, column=col_num).value,
                context=f"row {row}, {name}, {cat_name}",
            )
            for cat_name, col_num in category_cols.items()
        }
    return result


def _get_week_range():
    """Return (start, end) for the current week's report range: Monday through
    yesterday, with the Monday special-case (show the previous complete week).

    Shared by get_cumulative_data() (Weekly Performance Report) and
    get_weekly_daily_breakdown() (weekly change-detection snapshot) so the
    date-range calculation lives in one place.
    """
    today = datetime.now().date()
    weekday = today.weekday()  # Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6
    if weekday == 0:  # Monday → show previous complete week
        start = today - timedelta(days=7)   # previous Monday
    else:
        start = today - timedelta(days=weekday)  # this Monday
    end = today - timedelta(days=1)
    return start, end


def get_cumulative_data():
    """Determine date range and read cumulative data from Excel."""
    today = datetime.now().date()
    weekday = today.weekday()  # Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6

    start, end = _get_week_range()

    if weekday == 0:  # Monday → show previous complete week same as Sunday
        week_number = start.isocalendar()[1]
        week_label = f"Week #{week_number} - {start.strftime('%b %d')} to {end.strftime('%b %d')} (Complete Week)"
        prev_day_start = today - timedelta(days=2)  # previous Saturday (for Weekly Report card only)
        prev_day_end = today - timedelta(days=2)    # previous Saturday (for Weekly Report card only)
        breakdown_prev_day_start = today - timedelta(days=1)  # previous Sunday (for breakdown card — no data, will show 0s)
        breakdown_prev_day_end = today - timedelta(days=1)    # previous Sunday (for breakdown card — no data, will show 0s)
    elif weekday == 6:  # Sunday → this week Mon to yesterday (Saturday = complete week)
        week_number = start.isocalendar()[1]
        week_label = f"Week #{week_number} - {start.strftime('%b %d')} to {end.strftime('%b %d')} (Complete Week)"
        prev_day_start = end
        prev_day_end = end
        breakdown_prev_day_start = prev_day_start
        breakdown_prev_day_end = prev_day_end
    else:  # Tue to Sat → this Monday to yesterday
        week_number = start.isocalendar()[1]
        if start == end:
            week_label = f"Week #{week_number} - {start.strftime('%b %d')}"
        else:
            week_label = f"Week #{week_number} - {start.strftime('%b %d')} to {end.strftime('%b %d')}"
        prev_day_start = today - timedelta(days=1)
        prev_day_end = today - timedelta(days=1)
        breakdown_prev_day_start = prev_day_start
        breakdown_prev_day_end = prev_day_end

    # Load Excel
    wb = load_workbook(TEMP_FILE)
    if "Daily Performance Bonus" in wb.sheetnames:
        ws = wb["Daily Performance Bonus"]
    else:
        ws = max(wb.worksheets, key=lambda s: s.max_row or 0)

    # Detect columns from row 1
    date_col, name_col, category_cols, point_cols, notes_col = _detect_columns(ws)

    if not name_col:
        print(f"[ERROR] Could not find Name column in sheet '{ws.title}'.")
        return None, None, None

    # Get ALL unique employee names from entire sheet
    all_names = []
    seen = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=name_col).value
        if val and str(val).strip() and str(val).strip() not in seen:
            seen.add(str(val).strip())
            all_names.append(str(val).strip())

    # Aggregate points per employee within date range
    totals = defaultdict(int)
    prev_day_totals = defaultdict(int)
    for name in all_names:
        totals[name] = 0
        prev_day_totals[name] = 0

    # Deduplicate rows by (name, date) to prevent double-counting
    row_data = {}  # key: (name, date), value: points
    prev_day_breakdown = {}  # key: (name, date), value: {category: value}
    if date_col:
        for row in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row, column=name_col).value
            date_val = ws.cell(row=row, column=date_col).value
            if not name_val or not date_val:
                continue
            row_date = date_val.date() if hasattr(date_val, 'date') else None
            if not row_date:
                continue
            name = str(name_val).strip()
            pts = sum(
                _safe_int(ws.cell(row=row, column=c).value, context=f"row {row}, {name}")
                for c in point_cols
            )
            key = (name, row_date)
            if key in row_data:
                print(f"[WARNING] Duplicate entry found for {name} on {row_date} — using last occurrence, please check source file")
            row_data[key] = pts

            if breakdown_prev_day_start <= row_date <= breakdown_prev_day_end:
                category_values = {}
                for cat_name, col_num in category_cols.items():
                    val = ws.cell(row=row, column=col_num).value
                    category_values[cat_name] = _safe_int(val, context=f"row {row}, {name}, {cat_name}")
                notes_val = ws.cell(row=row, column=notes_col).value if notes_col else None
                category_values["Notes"] = str(notes_val).strip() if notes_val else ""
                prev_day_breakdown[key] = category_values

    # Calculate weekly and previous-day totals from deduplicated data
    for (name, row_date), pts in row_data.items():
        if start <= row_date <= end:
            totals[name] += pts
        if prev_day_start <= row_date <= prev_day_end:
            prev_day_totals[name] += pts

    employees = sorted(
        [{"name": n, "points": p, "amount": p * 10, "prev_day_points": prev_day_totals.get(n, 0), "prev_day_amount": prev_day_totals.get(n, 0) * 10} for n, p in totals.items()],
        key=lambda x: (-x["prev_day_amount"], x["name"])
    )

    prev_day_label = breakdown_prev_day_start.strftime('%b %d')
    prev_day_employees = []
    for (name, row_date), categories in prev_day_breakdown.items():
        prev_day_employees.append({
            "name": name,
            **categories
        })

    # For Monday: fill in all employees with 0s for categories (since Sunday has no data)
    if weekday == 0:
        names_in_breakdown = {emp["name"] for emp in prev_day_employees}
        for emp_name in all_names:
            if emp_name not in names_in_breakdown:
                zero_categories = {cat: 0 for cat in category_cols.keys()}
                zero_categories["Notes"] = ""
                prev_day_employees.append({
                    "name": emp_name,
                    **zero_categories
                })

    prev_day_employees.sort(key=lambda x: x["name"])

    prev_day_breakdown_data = {
        "date_label": prev_day_label,
        "employees": prev_day_employees
    }

    return week_label, employees, prev_day_breakdown_data


def get_weekly_daily_breakdown():
    """Read per-employee, per-day points for the current week's report range.

    Uses the same range as get_cumulative_data() (via _get_week_range()) and the
    same row-dedup-by-(name,date) and _safe_int() handling. Returns
    {date_iso: {employee_name: points}} — one entry per calendar day with data,
    used to detect day-level corrections within the current week's snapshot.
    """
    start, end = _get_week_range()

    wb = load_workbook(TEMP_FILE)
    if "Daily Performance Bonus" in wb.sheetnames:
        ws = wb["Daily Performance Bonus"]
    else:
        ws = max(wb.worksheets, key=lambda s: s.max_row or 0)

    date_col, name_col, _, point_cols, _ = _detect_columns(ws)
    if not name_col or not date_col:
        return {}

    row_data = {}  # key: (name, date), value: points
    for row in range(2, ws.max_row + 1):
        name_val = ws.cell(row=row, column=name_col).value
        date_val = ws.cell(row=row, column=date_col).value
        if not name_val or not date_val:
            continue
        row_date = date_val.date() if hasattr(date_val, "date") else None
        if not row_date:
            continue
        name = str(name_val).strip()
        pts = sum(
            _safe_int(ws.cell(row=row, column=c).value, context=f"row {row}, {name}")
            for c in point_cols
        )
        row_data[(name, row_date)] = pts

    result = {}
    for (name, row_date), pts in row_data.items():
        if start <= row_date <= end:
            result.setdefault(row_date.isoformat(), {})[name] = pts

    return result


def format_prev_day_card(prev_day_breakdown, title_prefix=""):
    """Build standalone Adaptive Card for previous day performance breakdown."""
    if not prev_day_breakdown or not prev_day_breakdown["employees"]:
        return None

    first_emp = prev_day_breakdown["employees"][0]
    has_notes = "Notes" in first_emp
    categories = [k for k in first_emp.keys() if k not in ("name", "Notes")]

    header_columns = [
        {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": "Name", "weight": "Bolder", "wrap": True}]}
    ] + [
        {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": cat, "weight": "Bolder", "horizontalAlignment": "Center", "wrap": True}]}
        for cat in categories
    ]
    if has_notes:
        header_columns.append(
            {"type": "Column", "width": "4", "items": [{"type": "TextBlock", "text": "Notes", "weight": "Bolder", "wrap": True}]}
        )

    body = [
        {"type": "TextBlock", "text": f"{title_prefix}Previous Day Performance Breakdown", "weight": "Bolder", "size": "Large", "color": "Accent", "wrap": True},
        {"type": "TextBlock", "text": f"Date: {prev_day_breakdown['date_label']}", "weight": "Normal", "size": "Medium", "spacing": "None", "wrap": True},
        {"type": "ColumnSet", "separator": True, "columns": header_columns}
    ]

    for emp in prev_day_breakdown["employees"]:
        row_columns = [
            {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": str(emp["name"]), "wrap": True}]}
        ] + [
            {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": str(emp.get(cat, 0)), "horizontalAlignment": "Center", "wrap": True}]}
            for cat in categories
        ]
        if has_notes:
            row_columns.append(
                {"type": "Column", "width": "4", "items": [{"type": "TextBlock", "text": str(emp.get("Notes", "")), "wrap": True}]}
            )
        body.append({"type": "ColumnSet", "separator": True, "style": "default", "columns": row_columns})

    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": body
    }


def format_teams_message(week_label, employees, title_prefix=""):
    body = [
        {"type": "TextBlock", "text": f"{title_prefix}Weekly Performance Report", "weight": "Bolder", "size": "Large", "color": "Accent", "wrap": True},
        {"type": "TextBlock", "text": week_label, "weight": "Bolder", "size": "Medium", "spacing": "None", "wrap": True},
        {
            "type": "ColumnSet",
            "separator": True,
            "columns": [
                {"type": "Column", "width": "3", "items": [{"type": "TextBlock", "text": "Name", "weight": "Bolder", "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": "Previous Day Point", "weight": "Bolder", "horizontalAlignment": "Center", "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": "Previous Day Amount", "weight": "Bolder", "horizontalAlignment": "Right", "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": "Weekly Total Point", "weight": "Bolder", "horizontalAlignment": "Center", "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": "Weekly Total Amount", "weight": "Bolder", "horizontalAlignment": "Right", "wrap": True}]}
            ]
        }
    ]

    for emp in employees:
        name = emp["name"]
        prev_day_points = int(emp["prev_day_points"]) if emp["prev_day_points"] is not None else 0
        prev_day_amount = int(emp["prev_day_amount"]) if emp["prev_day_amount"] is not None else 0
        weekly_points = int(emp["points"]) if emp["points"] is not None else 0
        weekly_amount = int(emp["amount"]) if emp["amount"] else 0
        body.append({
            "type": "ColumnSet",
            "separator": True,
            "style": "default",
            "columns": [
                {"type": "Column", "width": "3", "items": [{"type": "TextBlock", "text": str(name), "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": str(prev_day_points), "horizontalAlignment": "Center", "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": f"₹{prev_day_amount}", "horizontalAlignment": "Right", "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": str(weekly_points), "horizontalAlignment": "Center", "wrap": True}]},
                {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": f"₹{weekly_amount}", "horizontalAlignment": "Right", "wrap": True}]}
            ]
        })

    total_prev_day_points = sum(int(emp["prev_day_points"]) if emp["prev_day_points"] else 0 for emp in employees)
    total_prev_day_amount = total_prev_day_points * 10
    total_weekly_points = sum(int(emp["points"]) if emp["points"] else 0 for emp in employees)
    total_weekly_amount = total_weekly_points * 10
    body.append({
        "type": "ColumnSet",
        "separator": True,
        "columns": [
            {"type": "Column", "width": "3", "items": [{"type": "TextBlock", "text": "TOTAL", "weight": "Bolder", "wrap": True}]},
            {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": str(total_prev_day_points), "weight": "Bolder", "horizontalAlignment": "Center", "wrap": True}]},
            {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": f"₹{total_prev_day_amount}", "weight": "Bolder", "horizontalAlignment": "Right", "wrap": True}]},
            {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": str(total_weekly_points), "weight": "Bolder", "horizontalAlignment": "Center", "wrap": True}]},
            {"type": "Column", "width": "2", "items": [{"type": "TextBlock", "text": f"₹{total_weekly_amount}", "weight": "Bolder", "horizontalAlignment": "Right", "wrap": True}]}
        ]
    })

    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": body
    }


def send_teams_webhook_message(webhook_url, adaptive_card):
    headers = {"Content-Type": "application/json"}
    payload = {
        "type": "message",
        "attachments": [{"contentType": "application/vnd.microsoft.card.adaptive", "content": adaptive_card}]
    }
    try:
        response = requests.post(webhook_url, headers=headers, json=payload)
        if response.status_code in [200, 201, 202]:
            return True
        print(f"ERROR sending message: {response.status_code}")
        if response.text:
            print(f"Response: {response.text[:300]}")
        return False
    except Exception as e:
        print(f"ERROR: Could not send Teams message: {e}")
        return False


def cleanup_temp_file():
    for f in [TEMP_FILE]:
        try:
            if os.path.exists(f):
                os.remove(f)
                print(f"[OK] Deleted: {f}")
        except Exception as e:
            print(f"WARNING: Could not delete {f}: {e}")
    print("[OK] Cleanup done")
    return True


def load_snapshot():
    """Load the last-known yesterday snapshot from disk. Returns None if absent."""
    try:
        with open(SNAPSHOT_FILE, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        return None
    except Exception as e:
        print(f"[WARNING] Could not read snapshot: {e}")
        return None


def save_snapshot(date_str, employees_dict, week_start_str=None, week_days_dict=None, week_totals_dict=None):
    """Persist yesterday's per-employee category data as the new baseline.

    week_start_str/week_days_dict are optional so existing day-only callers
    (and tests) keep working unchanged; when provided, the current week's
    per-day breakdown is stored alongside under the "week" key.
    week_totals_dict is optional the same way; when provided, it stores the
    aggregated weekly Points/Amount per employee (see get_weekly_totals())
    under "week"."totals" for the authoritative Weekly Report total check.
    """
    try:
        payload = {"date": date_str, "employees": employees_dict}
        if week_start_str is not None:
            payload["week"] = {"week_start": week_start_str, "days": week_days_dict or {}}
            if week_totals_dict is not None:
                payload["week"]["totals"] = week_totals_dict
        with open(SNAPSHOT_FILE, "w") as f:
            json.dump(payload, f, indent=2)
        print(f"[OK] Snapshot saved for {date_str}")
    except Exception as e:
        print(f"[WARNING] Could not save snapshot: {e}")


def has_yesterday_data_changed(yesterday_date_str, fresh_data):
    """Return True only when yesterday's snapshot exists, date matches, and data differs."""
    snapshot = load_snapshot()
    if snapshot is None or snapshot.get("date") != yesterday_date_str:
        return False
    return snapshot["employees"] != fresh_data


def has_weekly_data_changed(week_start_str, fresh_days):
    """Return True only when a stored weekly snapshot exists for the same
    week_start and at least one day present in BOTH the stored and fresh data
    has different employee points (an actual correction).

    A missing snapshot, a missing "week" key (old-format snapshot file), or a
    different week_start (a new week has started) all return False — those
    are not corrections. A day present only in fresh_days (the week grew by
    one day since the last run) is expected growth and is ignored here too.

    NOTE: this is a proxy for "did the Weekly Report's totals change", not an
    exact match — it only visits days present in both stored and fresh data,
    so it cannot see an entire day's rows disappearing from the source file.
    has_weekly_totals_changed() below is the authoritative check for that;
    this function is kept as an additional (more granular, day-level) signal.
    """
    snapshot = load_snapshot()
    week_snapshot = snapshot.get("week") if snapshot else None
    if not week_snapshot or week_snapshot.get("week_start") != week_start_str:
        return False
    stored_days = week_snapshot.get("days", {})
    for day, employees in fresh_days.items():
        if day in stored_days and stored_days[day] != employees:
            return True
    return False


def get_weekly_totals(employees):
    """Extract {name: {"points":, "amount":}} from get_cumulative_data()'s
    employees list — the exact Weekly Total Point/Amount values shown on the
    Weekly Performance Report card, per employee."""
    return {emp["name"]: {"points": emp["points"], "amount": emp["amount"]} for emp in employees}


def has_weekly_totals_changed(week_start_str, fresh_totals):
    """Return True only when a stored weekly-totals snapshot exists for the
    same week_start and differs from fresh_totals.

    This is the authoritative Weekly Report check: it compares the actual
    aggregated weekly Points/Amount per employee (as shown on the card), so
    unlike has_weekly_data_changed() it also catches a whole day's rows
    disappearing from the source file, or any other change that shifts the
    total without necessarily changing an individual day's record.

    A missing snapshot, missing "week"/"totals" key (old-format snapshot), or
    a different week_start (a new week has started) all return False — those
    are not corrections.
    """
    snapshot = load_snapshot()
    week_snapshot = snapshot.get("week") if snapshot else None
    if not week_snapshot or week_snapshot.get("week_start") != week_start_str:
        return False
    stored_totals = week_snapshot.get("totals")
    if stored_totals is None:
        return False
    return stored_totals != fresh_totals


# MAIN EXECUTION
if __name__ == "__main__":
    print("=" * 60)
    print("WEEKLY PERFORMANCE REPORT - TEAMS SENDER")
    print("=" * 60)

    if not download_source_file():
        exit(1)

    yesterday_date = datetime.now().date() - timedelta(days=1)
    yesterday_str = yesterday_date.isoformat()
    fresh_yesterday_data = get_real_yesterday_data()

    week_start, _ = _get_week_range()
    week_start_str = week_start.isoformat()
    fresh_week_days = get_weekly_daily_breakdown()

    print("Determining date range and reading Excel data...")
    week_label, employee_data, prev_day_breakdown = get_cumulative_data()

    if not week_label or not employee_data:
        print("[ERROR] No data to send.")
        cleanup_temp_file()
        exit(1)

    fresh_weekly_totals = get_weekly_totals(employee_data)

    snapshot = load_snapshot()

    # Condition 1: date check — does yesterday's date match the snapshot's stored date.
    if snapshot is None:
        date_changed = True
        daily_correction = False
    elif snapshot["date"] != yesterday_str:
        date_changed = True
        daily_correction = False
    # Condition 2: Previous Day Performance Breakdown check — per-employee,
    # per-category data for yesterday vs. what's stored in the snapshot.
    elif snapshot["employees"] != fresh_yesterday_data:
        date_changed = False
        daily_correction = True
    else:
        date_changed = False
        daily_correction = False

    # Condition 3: Weekly Performance Report check — aggregated weekly
    # Points/Amount per employee (as shown on the card), vs. what's stored.
    # weekly_totals_correction is the authoritative check; weekly_day_correction
    # is kept as an additional day-level signal (see has_weekly_data_changed()).
    weekly_day_correction = has_weekly_data_changed(week_start_str, fresh_week_days)
    weekly_totals_correction = has_weekly_totals_changed(week_start_str, fresh_weekly_totals)
    weekly_correction = weekly_day_correction or weekly_totals_correction

    changed = date_changed or daily_correction or weekly_correction
    is_correction = daily_correction or weekly_correction

    print("\nChange detection:")
    print(f"  [1] Date changed (yesterday vs. snapshot date): {date_changed}")
    print(f"  [2] Previous Day Performance Breakdown changed (per-category, yesterday): {daily_correction}")
    print(f"  [3] Weekly Performance Report total changed (per-employee weekly points/amount): {weekly_totals_correction}"
          f"{' [also flagged by raw per-day breakdown]' if weekly_day_correction and not weekly_totals_correction else ''}")

    if not changed:
        save_snapshot(yesterday_str, fresh_yesterday_data, week_start_str, fresh_week_days, fresh_weekly_totals)
        print("[INFO] No change detected since last check (date, Previous Day Breakdown, and Weekly Report totals all match). Exiting silently.")
        print("[INFO] Baseline snapshot updated (no send needed)")
        cleanup_temp_file()
        exit(0)

    reasons = []
    if date_changed:
        reasons.append("date changed")
    if daily_correction:
        reasons.append("Previous Day Performance Breakdown data changed")
    if weekly_correction:
        reasons.append("Weekly Performance Report total changed")
    print(f"[INFO] Sending — triggered by: {', '.join(reasons)}")

    title_prefix = "🔧 Corrected Report - " if is_correction else ""

    print(f"[OK] Report: {week_label}")
    print(f"[OK] Employees: {len(employee_data)}")
    if prev_day_breakdown and prev_day_breakdown["employees"]:
        print(f"[OK] Previous Day ({prev_day_breakdown['date_label']}): {len(prev_day_breakdown['employees'])} employees")
    print("\nVerification:")
    for emp in employee_data:
        print(f"  {emp['name']}: Prev Day {emp['prev_day_points']} pts (₹{emp['prev_day_amount']}) | Weekly {emp['points']} pts (₹{emp['amount']})")
    if prev_day_breakdown and prev_day_breakdown["employees"]:
        print(f"\nPrevious Day Breakdown ({prev_day_breakdown['date_label']}):")
        for emp in prev_day_breakdown["employees"]:
            print(
                f"  {emp['name']}: Punctuality={emp.get('Punctuality', 0)}, "
                f"L&D={emp.get('L&D', 0)}, Fluency Compliance={emp.get('Fluency Compliance', 0)}, "
                f"Innovation={emp.get('Innovation', 0)}, "
                f"Extraordinary Performance={emp.get('Extraordinary Performance', 0)}"
            )
    print()

    cards_sent = 0

    if prev_day_breakdown and prev_day_breakdown["employees"]:
        prev_day_card = format_prev_day_card(prev_day_breakdown, title_prefix=title_prefix)
        if prev_day_card:
            print("Sending Previous Day Performance Breakdown card to Teams...")
            if send_teams_webhook_message(WEBHOOK_URL, prev_day_card):
                print("[OK] Previous Day card sent successfully!")
                cards_sent += 1
            else:
                print("[ERROR] Failed to send Previous Day card to Teams")
                cleanup_temp_file()
                exit(1)

    weekly_card = format_teams_message(week_label, employee_data, title_prefix=title_prefix)
    print("Sending Weekly Performance Report card to Teams...")
    if send_teams_webhook_message(WEBHOOK_URL, weekly_card):
        print("[OK] Weekly Report card sent successfully!")
        cards_sent += 1
    else:
        print("[ERROR] Failed to send Weekly Report card to Teams")
        cleanup_temp_file()
        exit(1)

    save_snapshot(yesterday_str, fresh_yesterday_data, week_start_str, fresh_week_days, fresh_weekly_totals)
    print(f"\n[OK] Total cards sent: {cards_sent}")
    cleanup_temp_file()
    print("=" * 60)

