import json
import os
import sys
from datetime import date, datetime, timedelta
from unittest.mock import patch

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
import weekly_report_send_teams as module

CATEGORIES = [
    "Punctuality", "L&D", "Fluency Compliance",
    "Innovation", "Extraordinary Performance",
]


def make_test_workbook(rows):
    """Build an in-memory openpyxl Workbook with the expected sheet structure.

    rows: list of dicts with keys: date (datetime), name (str),
          and optionally each category name (int, default 0).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Performance Bonus"
    headers = ["Index", "Date", "Name"] + CATEGORIES
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=row["date"])
        ws.cell(row=r, column=3, value=row["name"])
        for col, cat in enumerate(CATEGORIES, 4):
            ws.cell(row=r, column=col, value=row.get(cat, 0))
    return wb


# ── _detect_columns ──────────────────────────────────────────────────────────

def test_detect_columns_identifies_all_columns():
    wb = make_test_workbook([])
    ws = wb["Daily Performance Bonus"]
    date_col, name_col, category_cols, point_cols, notes_col = module._detect_columns(ws)
    assert date_col == 2
    assert name_col == 3
    assert category_cols == {
        "Punctuality": 4, "L&D": 5, "Fluency Compliance": 6,
        "Innovation": 7, "Extraordinary Performance": 8,
    }
    assert sorted(point_cols) == [4, 5, 6, 7, 8]
    assert notes_col is None


def test_detect_columns_returns_none_for_missing_columns():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Index")
    date_col, name_col, category_cols, point_cols, notes_col = module._detect_columns(ws)
    assert date_col is None
    assert name_col is None
    assert notes_col is None


# ── get_real_yesterday_data ───────────────────────────────────────────────────

def test_get_real_yesterday_data_returns_only_yesterday(tmp_path):
    yesterday = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    today_dt = datetime.combine(date.today(), datetime.min.time())

    wb = make_test_workbook([
        {"date": yesterday, "name": "Alice",
         "Punctuality": 1, "L&D": 1, "Fluency Compliance": 0,
         "Innovation": 0, "Extraordinary Performance": 0},
        {"date": today_dt, "name": "Alice",
         "Punctuality": 1, "L&D": 1, "Fluency Compliance": 1,
         "Innovation": 1, "Extraordinary Performance": 1},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_real_yesterday_data()

    assert list(result.keys()) == ["Alice"]
    assert result["Alice"] == {
        "Punctuality": 1, "L&D": 1, "Fluency Compliance": 0,
        "Innovation": 0, "Extraordinary Performance": 0,
    }


def test_get_real_yesterday_data_deduplicates_keeps_last(tmp_path):
    yesterday = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())

    wb = make_test_workbook([
        {"date": yesterday, "name": "Bob", "Punctuality": 0, "L&D": 0,
         "Fluency Compliance": 0, "Innovation": 0, "Extraordinary Performance": 0},
        {"date": yesterday, "name": "Bob", "Punctuality": 1, "L&D": 1,
         "Fluency Compliance": 1, "Innovation": 1, "Extraordinary Performance": 1},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_real_yesterday_data()

    assert result["Bob"]["Punctuality"] == 1


def test_get_real_yesterday_data_empty_when_no_rows(tmp_path):
    wb = make_test_workbook([])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_real_yesterday_data()

    assert result == {}


# ── Snapshot functions ────────────────────────────────────────────────────────

def test_save_and_load_roundtrip(tmp_path):
    data = {"Alice": {"Punctuality": 1, "L&D": 0, "Fluency Compliance": 1,
                      "Innovation": 0, "Extraordinary Performance": 0}}
    snap_file = str(tmp_path / "snap.json")
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        module.save_snapshot("2026-06-24", data)
        result = module.load_snapshot()
    assert result == {"date": "2026-06-24", "employees": data}


def test_load_snapshot_returns_none_when_file_missing(tmp_path):
    with patch.object(module, "SNAPSHOT_FILE", str(tmp_path / "missing.json")):
        assert module.load_snapshot() is None


def test_save_snapshot_silent_on_bad_path():
    with patch.object(module, "SNAPSHOT_FILE", "/nonexistent_dir/snap.json"):
        module.save_snapshot("2026-06-24", {})  # must not raise


# ── has_yesterday_data_changed ────────────────────────────────────────────────

def test_no_change_when_data_identical(tmp_path):
    data = {"Alice": {"Punctuality": 1, "L&D": 0, "Fluency Compliance": 1,
                      "Innovation": 0, "Extraordinary Performance": 0}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-24", "employees": data}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert not module.has_yesterday_data_changed("2026-06-24", data)


def test_change_detected_when_value_differs(tmp_path):
    old = {"Alice": {"Punctuality": 0, "L&D": 0, "Fluency Compliance": 0,
                     "Innovation": 0, "Extraordinary Performance": 0}}
    new = {"Alice": {"Punctuality": 1, "L&D": 0, "Fluency Compliance": 0,
                     "Innovation": 0, "Extraordinary Performance": 0}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-24", "employees": old}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert module.has_yesterday_data_changed("2026-06-24", new)


def test_no_change_when_snapshot_date_differs(tmp_path):
    snapshot_data = {"Alice": {"Punctuality": 1, "L&D": 0, "Fluency Compliance": 0,
                               "Innovation": 0, "Extraordinary Performance": 0}}
    fresh_data = {"Alice": {"Punctuality": 0, "L&D": 0, "Fluency Compliance": 0,
                            "Innovation": 0, "Extraordinary Performance": 0}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-23", "employees": snapshot_data}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        # Different date means watcher fired on a new day — not a correction
        assert not module.has_yesterday_data_changed("2026-06-24", fresh_data)


def test_no_change_when_no_snapshot(tmp_path):
    with patch.object(module, "SNAPSHOT_FILE", str(tmp_path / "missing.json")):
        assert not module.has_yesterday_data_changed("2026-06-24", {"Alice": {}})


def test_baseline_established_without_send_enables_future_correction_detection(tmp_path):
    """Regression test: save_snapshot() must persist a baseline even on a run
    where nothing changed (so no Teams send happened), otherwise a "week" key
    never gets written and weekly corrections can never be detected on any
    later run. Simulates two runs: the first establishes the baseline (as the
    fixed __main__ now does on its changed=False, save-anyway path), the
    second shows a later edit to that same baseline data is now caught."""
    snap_file = str(tmp_path / "snap.json")
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert module.load_snapshot() is None  # no prior snapshot at all

        yesterday_employees = {"Alice": {"Punctuality": 1, "L&D": 0, "Fluency Compliance": 0,
                                          "Innovation": 0, "Extraordinary Performance": 0}}
        week_days = {"2026-06-22": {"Alice": 2}, "2026-06-23": {"Alice": 3}}
        module.save_snapshot("2026-06-24", yesterday_employees, "2026-06-22", week_days)

        snapshot = module.load_snapshot()
        assert snapshot["week"] == {"week_start": "2026-06-22", "days": week_days}

        # Next run: yesterday itself is unchanged, but Monday's already-recorded
        # data was corrected in the source file — must now be detectable because
        # a baseline exists.
        corrected_week_days = {"2026-06-22": {"Alice": 5}, "2026-06-23": {"Alice": 3}}
        assert not module.has_yesterday_data_changed("2026-06-24", yesterday_employees)
        assert module.has_weekly_data_changed("2026-06-22", corrected_week_days)


# ── Dynamic column detection ──────────────────────────────────────────────────

CATEGORIES_PLUS = CATEGORIES + ["Teamwork"]


def make_test_workbook_extra(rows):
    """Like make_test_workbook but with an extra 'Teamwork' column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Performance Bonus"
    headers = ["Index", "Date", "Name"] + CATEGORIES_PLUS
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=row["date"])
        ws.cell(row=r, column=3, value=row["name"])
        for col, cat in enumerate(CATEGORIES_PLUS, 4):
            ws.cell(row=r, column=col, value=row.get(cat, 0))
    return wb


def test_detect_columns_includes_extra_column():
    wb = make_test_workbook_extra([])
    ws = wb["Daily Performance Bonus"]
    date_col, name_col, category_cols, point_cols, notes_col = module._detect_columns(ws)
    assert "Teamwork" in category_cols
    assert category_cols["Teamwork"] == 9  # 9th column: Index, Date, Name + 5 + Teamwork
    assert sorted(point_cols) == [4, 5, 6, 7, 8, 9]


def test_get_real_yesterday_data_includes_extra_column(tmp_path):
    yesterday = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook_extra([
        {"date": yesterday, "name": "Carol",
         "Punctuality": 1, "L&D": 0, "Fluency Compliance": 1,
         "Innovation": 0, "Extraordinary Performance": 0, "Teamwork": 1},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_real_yesterday_data()

    assert "Carol" in result
    assert result["Carol"]["Teamwork"] == 1


# ── Notes / free-text column exclusion ────────────────────────────────────────

def make_test_workbook_with_notes(rows):
    """Like make_test_workbook but with an extra free-text 'Notes' column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Performance Bonus"
    headers = ["Index", "Date", "Name"] + CATEGORIES + ["Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=row["date"])
        ws.cell(row=r, column=3, value=row["name"])
        for col, cat in enumerate(CATEGORIES, 4):
            ws.cell(row=r, column=col, value=row.get(cat, 0))
        ws.cell(row=r, column=len(headers), value=row.get("Notes", ""))
    return wb


def test_detect_columns_excludes_notes_column():
    wb = make_test_workbook_with_notes([])
    ws = wb["Daily Performance Bonus"]
    date_col, name_col, category_cols, point_cols, notes_col = module._detect_columns(ws)
    assert "Notes" not in category_cols
    assert sorted(point_cols) == [4, 5, 6, 7, 8]  # Notes column (9) excluded
    assert notes_col == 9


def test_detect_columns_excludes_notes_case_insensitive():
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Performance Bonus"
    headers = ["Index", "Date", "Name"] + CATEGORIES + ["NOTES"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    date_col, name_col, category_cols, point_cols, notes_col = module._detect_columns(ws)
    assert "NOTES" not in category_cols
    assert sorted(point_cols) == [4, 5, 6, 7, 8]
    assert notes_col == 9


def test_get_cumulative_data_ignores_notes_column_text(tmp_path):
    """A free-text Notes column with a value like 'Washroom work+ Ticket' must
    not crash get_cumulative_data — it should simply be excluded as a point column."""
    yesterday_dt = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook_with_notes([
        {"date": yesterday_dt, "name": "Dave",
         "Punctuality": 1, "L&D": 1, "Fluency Compliance": 1,
         "Innovation": 0, "Extraordinary Performance": 0,
         "Notes": "Washroom work+ Ticket"},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        week_label, employees, prev_day_breakdown = module.get_cumulative_data()

    assert week_label is not None
    dave = next(e for e in employees if e["name"] == "Dave")
    assert dave["points"] == 3


def test_get_real_yesterday_data_ignores_notes_column_text(tmp_path):
    """A free-text Notes column must not crash get_real_yesterday_data, and the
    returned per-category dict must not include Notes (numeric-only snapshot)."""
    yesterday_dt = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook_with_notes([
        {"date": yesterday_dt, "name": "Gail",
         "Punctuality": 1, "L&D": 0, "Fluency Compliance": 0,
         "Innovation": 0, "Extraordinary Performance": 0,
         "Notes": "Washroom work+ Ticket"},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_real_yesterday_data()

    assert "Notes" not in result["Gail"]
    assert result["Gail"]["Punctuality"] == 1


def test_get_cumulative_data_notes_text_included_in_prev_day_breakdown(tmp_path):
    """Notes text should be captured verbatim per employee in the previous-day
    breakdown, not summed or dropped."""
    yesterday_dt = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook_with_notes([
        {"date": yesterday_dt, "name": "Heidi",
         "Punctuality": 1, "L&D": 1, "Fluency Compliance": 1,
         "Innovation": 0, "Extraordinary Performance": 0,
         "Notes": "Washroom work+ Ticket"},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        week_label, employees, prev_day_breakdown = module.get_cumulative_data()

    heidi = next(e for e in prev_day_breakdown["employees"] if e["name"] == "Heidi")
    assert heidi["Notes"] == "Washroom work+ Ticket"


# ── _safe_int defense-in-depth ────────────────────────────────────────────────

def test_safe_int_returns_zero_for_none_and_empty():
    assert module._safe_int(None) == 0
    assert module._safe_int("") == 0


def test_safe_int_converts_valid_numbers():
    assert module._safe_int(5) == 5
    assert module._safe_int("3") == 3


def test_safe_int_returns_zero_and_warns_on_bad_value(capsys):
    result = module._safe_int("Washroom work+ Ticket", context="row 2, Dave, Punctuality")
    assert result == 0
    captured = capsys.readouterr()
    assert "[WARNING]" in captured.out
    assert "Washroom work+ Ticket" in captured.out
    assert "row 2, Dave, Punctuality" in captured.out


def test_get_cumulative_data_stray_text_in_point_column_treated_as_zero(capsys, tmp_path):
    """If a point column somehow contains stray text (not caught by column
    exclusion), the script must not crash and should treat it as 0 with a warning."""
    yesterday_dt = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook([
        {"date": yesterday_dt, "name": "Erin",
         "Punctuality": 1, "L&D": 1, "Fluency Compliance": 0,
         "Innovation": 0, "Extraordinary Performance": 0},
    ])
    ws = wb["Daily Performance Bonus"]
    # Corrupt a point column cell with free text, simulating misentered data.
    ws.cell(row=2, column=6, value="Washroom work+ Ticket")  # Fluency Compliance column
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        week_label, employees, prev_day_breakdown = module.get_cumulative_data()

    erin = next(e for e in employees if e["name"] == "Erin")
    assert erin["points"] == 2  # Punctuality(1) + L&D(1) + Fluency Compliance(treated as 0)

    captured = capsys.readouterr()
    assert "[WARNING]" in captured.out
    assert "Washroom work+ Ticket" in captured.out


def test_get_real_yesterday_data_stray_text_treated_as_zero(capsys, tmp_path):
    yesterday = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook([
        {"date": yesterday, "name": "Frank",
         "Punctuality": 1, "L&D": 0, "Fluency Compliance": 0,
         "Innovation": 0, "Extraordinary Performance": 0},
    ])
    ws = wb["Daily Performance Bonus"]
    ws.cell(row=2, column=5, value="stray text")  # L&D column
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_real_yesterday_data()

    assert result["Frank"]["L&D"] == 0
    captured = capsys.readouterr()
    assert "[WARNING]" in captured.out
    assert "stray text" in captured.out


# ── _get_week_range / get_weekly_daily_breakdown ──────────────────────────────

def test_get_week_range_ends_yesterday():
    start, end = module._get_week_range()
    assert end == date.today() - timedelta(days=1)
    assert start <= end


def test_get_weekly_daily_breakdown_returns_per_day_points(tmp_path):
    yesterday = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook([
        {"date": yesterday, "name": "Alice", "Punctuality": 1, "L&D": 1,
         "Fluency Compliance": 0, "Innovation": 0, "Extraordinary Performance": 0},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_weekly_daily_breakdown()

    yesterday_iso = (date.today() - timedelta(days=1)).isoformat()
    assert result[yesterday_iso] == {"Alice": 2}


def test_get_weekly_daily_breakdown_dedupes_by_name_and_date(tmp_path):
    yesterday = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
    wb = make_test_workbook([
        {"date": yesterday, "name": "Bob", "Punctuality": 0, "L&D": 0,
         "Fluency Compliance": 0, "Innovation": 0, "Extraordinary Performance": 0},
        {"date": yesterday, "name": "Bob", "Punctuality": 1, "L&D": 1,
         "Fluency Compliance": 1, "Innovation": 1, "Extraordinary Performance": 1},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_weekly_daily_breakdown()

    yesterday_iso = (date.today() - timedelta(days=1)).isoformat()
    assert result[yesterday_iso]["Bob"] == 5  # last occurrence wins


def test_get_weekly_daily_breakdown_excludes_days_outside_range(tmp_path):
    far_past = datetime.combine(date.today() - timedelta(days=30), datetime.min.time())
    wb = make_test_workbook([
        {"date": far_past, "name": "Carol", "Punctuality": 1, "L&D": 0,
         "Fluency Compliance": 0, "Innovation": 0, "Extraordinary Performance": 0},
    ])
    test_file = str(tmp_path / "source.xlsx")
    wb.save(test_file)

    with patch.object(module, "TEMP_FILE", test_file):
        result = module.get_weekly_daily_breakdown()

    assert result == {}


# ── save_snapshot / load_snapshot weekly round-trip ────────────────────────────

def test_save_and_load_weekly_roundtrip(tmp_path):
    employees = {"Alice": {"Punctuality": 1, "L&D": 0, "Fluency Compliance": 0,
                           "Innovation": 0, "Extraordinary Performance": 0}}
    week_days = {"2026-06-22": {"Alice": 2}, "2026-06-23": {"Alice": 3}}
    snap_file = str(tmp_path / "snap.json")
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        module.save_snapshot("2026-06-24", employees, "2026-06-22", week_days)
        result = module.load_snapshot()
    assert result["week"] == {"week_start": "2026-06-22", "days": week_days}


# ── has_weekly_data_changed ─────────────────────────────────────────────────────

def test_no_weekly_change_when_days_identical(tmp_path):
    week_days = {"2026-06-22": {"Alice": 2}, "2026-06-23": {"Alice": 3}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-23", "employees": {},
                   "week": {"week_start": "2026-06-22", "days": week_days}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert not module.has_weekly_data_changed("2026-06-22", week_days)


def test_weekly_correction_detected_when_past_day_changes(tmp_path):
    stored_days = {"2026-06-22": {"Alice": 2}, "2026-06-23": {"Alice": 3}}
    fresh_days = {"2026-06-22": {"Alice": 5}, "2026-06-23": {"Alice": 3}}  # Monday corrected
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-23", "employees": {},
                   "week": {"week_start": "2026-06-22", "days": stored_days}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert module.has_weekly_data_changed("2026-06-22", fresh_days)


def test_weekly_no_false_positive_when_new_day_added(tmp_path):
    stored_days = {"2026-06-22": {"Alice": 2}}
    fresh_days = {"2026-06-22": {"Alice": 2}, "2026-06-23": {"Alice": 3}}  # grew by one day
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-22", "employees": {},
                   "week": {"week_start": "2026-06-22", "days": stored_days}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert not module.has_weekly_data_changed("2026-06-22", fresh_days)


def test_weekly_new_week_resets_without_flagging_correction(tmp_path):
    stored_days = {"2026-06-15": {"Alice": 9}}
    fresh_days = {"2026-06-22": {"Alice": 1}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-19", "employees": {},
                   "week": {"week_start": "2026-06-15", "days": stored_days}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert not module.has_weekly_data_changed("2026-06-22", fresh_days)


def test_weekly_missing_week_key_in_old_snapshot_does_not_crash(tmp_path):
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-24", "employees": {}}, f)  # old-format, no "week" key
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        result = module.load_snapshot()
        assert "week" not in result
        assert not module.has_weekly_data_changed("2026-06-22", {"2026-06-22": {"Alice": 1}})


# ── get_weekly_totals / has_weekly_totals_changed ──────────────────────────────

def test_get_weekly_totals_extracts_points_and_amount():
    employees = [
        {"name": "Alice", "points": 5, "amount": 50, "prev_day_points": 2, "prev_day_amount": 20},
        {"name": "Bob", "points": 3, "amount": 30, "prev_day_points": 1, "prev_day_amount": 10},
    ]
    assert module.get_weekly_totals(employees) == {
        "Alice": {"points": 5, "amount": 50},
        "Bob": {"points": 3, "amount": 30},
    }


def test_no_weekly_totals_change_when_identical(tmp_path):
    totals = {"Alice": {"points": 5, "amount": 50}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-23", "employees": {},
                   "week": {"week_start": "2026-06-22", "days": {}, "totals": totals}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert not module.has_weekly_totals_changed("2026-06-22", totals)


def test_weekly_totals_change_detected_when_amount_differs(tmp_path):
    stored_totals = {"Alice": {"points": 5, "amount": 50}}
    fresh_totals = {"Alice": {"points": 6, "amount": 60}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-23", "employees": {},
                   "week": {"week_start": "2026-06-22", "days": {}, "totals": stored_totals}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert module.has_weekly_totals_changed("2026-06-22", fresh_totals)


def test_weekly_totals_no_change_on_missing_totals_key_old_snapshot(tmp_path):
    """Old-format snapshot with a "week" key but no "totals" sub-key must not crash
    and must not be treated as a correction (no baseline to compare against yet)."""
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-23", "employees": {},
                   "week": {"week_start": "2026-06-22", "days": {}}}, f)  # no "totals"
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert not module.has_weekly_totals_changed("2026-06-22", {"Alice": {"points": 1, "amount": 10}})


def test_weekly_totals_no_change_on_new_week(tmp_path):
    stored_totals = {"Alice": {"points": 9, "amount": 90}}
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-19", "employees": {},
                   "week": {"week_start": "2026-06-15", "days": {}, "totals": stored_totals}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        assert not module.has_weekly_totals_changed("2026-06-22", {"Alice": {"points": 1, "amount": 10}})


def test_weekly_totals_catches_whole_day_removed_when_per_day_check_misses_it(tmp_path):
    """Regression test for the gap in has_weekly_data_changed(): if an entire
    day's rows disappear from the source file, the per-day check never visits
    that day (its loop only iterates fresh_days) and misses the resulting drop
    in the weekly total. has_weekly_totals_changed(), comparing the actual
    aggregated total, must catch it."""
    stored_days = {"2026-06-22": {"Alice": 2}, "2026-06-23": {"Alice": 3}}
    fresh_days = {"2026-06-23": {"Alice": 3}}  # 2026-06-22 entirely gone
    stored_totals = {"Alice": {"points": 5, "amount": 50}}  # 2+3
    fresh_totals = {"Alice": {"points": 3, "amount": 30}}   # only 3 remains
    snap_file = str(tmp_path / "snap.json")
    with open(snap_file, "w") as f:
        json.dump({"date": "2026-06-23", "employees": {},
                   "week": {"week_start": "2026-06-22", "days": stored_days, "totals": stored_totals}}, f)
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        # The per-day proxy misses it entirely:
        assert not module.has_weekly_data_changed("2026-06-22", fresh_days)
        # The authoritative totals check catches it:
        assert module.has_weekly_totals_changed("2026-06-22", fresh_totals)


def test_save_and_load_weekly_totals_roundtrip(tmp_path):
    employees = {"Alice": {"Punctuality": 1, "L&D": 0, "Fluency Compliance": 0,
                           "Innovation": 0, "Extraordinary Performance": 0}}
    week_days = {"2026-06-22": {"Alice": 2}, "2026-06-23": {"Alice": 3}}
    week_totals = {"Alice": {"points": 5, "amount": 50}}
    snap_file = str(tmp_path / "snap.json")
    with patch.object(module, "SNAPSHOT_FILE", snap_file):
        module.save_snapshot("2026-06-24", employees, "2026-06-22", week_days, week_totals)
        result = module.load_snapshot()
    assert result["week"] == {"week_start": "2026-06-22", "days": week_days, "totals": week_totals}


def test_format_prev_day_card_includes_extra_category():
    prev_day_breakdown = {
        "date_label": "Jun 26",
        "employees": [
            {"name": "Alice", "Punctuality": 1, "L&D": 0, "Fluency Compliance": 1,
             "Innovation": 0, "Extraordinary Performance": 0, "Teamwork": 1},
        ]
    }
    card = module.format_prev_day_card(prev_day_breakdown)
    assert card is not None

    # The header ColumnSet is the third element in body (index 2)
    header_columnset = card["body"][2]
    header_texts = [
        col["items"][0]["text"]
        for col in header_columnset["columns"]
    ]
    assert "Teamwork" in header_texts


def test_format_prev_day_card_renders_notes_column():
    prev_day_breakdown = {
        "date_label": "Jun 26",
        "employees": [
            {"name": "Alice", "Punctuality": 1, "L&D": 0, "Fluency Compliance": 1,
             "Innovation": 0, "Extraordinary Performance": 0,
             "Notes": "Washroom work+ Ticket"},
        ]
    }
    card = module.format_prev_day_card(prev_day_breakdown)
    assert card is not None

    header_columnset = card["body"][2]
    header_columns = header_columnset["columns"]
    header_texts = [col["items"][0]["text"] for col in header_columns]
    assert header_texts[-1] == "Notes"

    notes_header_col = header_columns[-1]
    assert "horizontalAlignment" not in notes_header_col["items"][0]

    row_columnset = card["body"][3]
    row_columns = row_columnset["columns"]
    notes_col = row_columns[-1]
    notes_item = notes_col["items"][0]
    assert notes_item["text"] == "Washroom work+ Ticket"
    assert notes_item["wrap"] is True
    assert "horizontalAlignment" not in notes_item
